#!/usr/bin/env python3
"""
Скрипт для конвертации CSV в XLSX формат
Сохраняет логику генерации данных из Pascal legacy
"""

import csv
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_file, xlsx_file=None):
    """Конвертирует CSV файл в XLSX формат с подстановкой значений даты и времени"""
    
    if xlsx_file is None:
        xlsx_file = csv_file.replace('.csv', '.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetry Data"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    date_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader)
        
        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Записываем данные
        row_num = 2
        for row in reader:
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                
                # Форматирование для разных типов данных
                header = headers[col_num - 1].lower()
                
                # Timestamp - форматируем как дату/время
                if 'recorded_at' in header or '_at' in header:
                    try:
                        # Парсим ISO 8601 формат
                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        cell.value = dt
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                        cell.fill = date_fill
                    except:
                        pass
                
                # Логические блоки - форматируем
                elif 'is_active' in header or 'status' in header:
                    if value in ['ИСТИНА', 'true', '1']:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value in ['ЛОЖЬ', 'false', '0']:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Числа - форматируем как числа
                elif 'voltage' in header or 'temp' in header:
                    try:
                        cell.value = float(value)
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right')
                    except:
                        pass
                
                # Строки - выравнивание по левому краю
                else:
                    cell.alignment = Alignment(horizontal='left')
            
            row_num += 1
    
    # Автоматическая ширина столбцов
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(header)
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(xlsx_file)
    print(f"Конвертировано: {csv_file} -> {xlsx_file}")
    return xlsx_file

if __name__ == 'main':
    if len(sys.argv) < 2:
        print("Использование: python3 csv_to_xlsx.py <csv_file> [xlsx_file]")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    xlsx_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(csv_file):
        print(f"Файл не найден: {csv_file}")
        sys.exit(1)
    
    csv_to_xlsx(csv_file, xlsx_file)